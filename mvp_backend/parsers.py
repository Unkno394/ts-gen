from __future__ import annotations

import csv
import json
import logging
from copy import deepcopy
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from models import (
    FormDocumentModel,
    FormFieldResolution,
    LayoutLine,
    LayoutToken,
    OptionItem,
    ParsedFile,
    ParsedKvPair,
    ParsedSection,
    ParsedSheet,
    ParsedTextBlock,
    QuestionGroup,
    ScalarFieldCandidate,
    SourceCandidate,
    TargetField,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for Excel support
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency for legacy XLS support
    xlrd = None

PARSER_DIR = Path(__file__).resolve().parent / 'parser'
if str(PARSER_DIR) not in sys.path:
    sys.path.insert(0, str(PARSER_DIR))

from candidate_normalizer import build_candidate_source, build_source_candidates  # noqa: E402
from document_parser import parse_document  # noqa: E402
from form_layout import resolve_business_form_fields  # noqa: E402

PREVIEW_ROW_LIMIT = 5
SPARSE_ROW_RATIO_THRESHOLD = 0.35
IMAGE_FILE_TYPES = {'png', 'jpg', 'jpeg', 'bmp', 'gif', 'tif', 'tiff', 'webp'}
PDF_FORM_ZONE_MIN_CONFIDENCE = 0.65
PDF_TABLE_ZONE_MIN_CONFIDENCE = 0.7
logger = logging.getLogger(__name__)

FORM_CRITICAL_TARGET_FIELDS = {
    'isresidentrf',
    'istaxresidencyonlyrf',
    'fatcabeneficiaryoptionlist',
}


class ParseError(ValueError):
    pass


def parse_file(file_path: Path, original_name: str | None = None) -> ParsedFile:
    ext = file_path.suffix.lower().lstrip('.')
    original_name = original_name or file_path.name
    parsed_kwargs: dict[str, Any] = {
        'columns': [],
        'rows': [],
        'content_type': 'unknown',
        'document_mode': 'data_table_mode',
        'extraction_status': 'unknown',
        'raw_text': '',
        'text_blocks': [],
        'sections': [],
        'kv_pairs': [],
        'source_candidates': [],
        'sheets': [],
        'form_model': None,
        'pdf_zone_summary': {},
        'warnings': [],
    }

    logger.info('parse_file started: name=%s ext=%s path=%s', original_name, ext, file_path)

    if ext == 'csv':
        columns, rows = _parse_csv(file_path)
        parsed_kwargs.update(
            {
                'columns': columns,
                'rows': rows[:PREVIEW_ROW_LIMIT],
                'content_type': 'table',
                'extraction_status': 'structured_extracted',
                'source_candidates': [
                    SourceCandidate(**candidate)
                    for candidate in build_source_candidates(
                        tables=[{'name': original_name, 'columns': columns, 'rows': rows[:PREVIEW_ROW_LIMIT]}]
                    )
                ],
            }
        )
    elif ext in {'xlsx', 'xls'}:
        columns, rows, extra_warnings, sheets = _parse_excel(file_path)
        parsed_kwargs.update(
            {
                'columns': columns,
                'rows': rows[:PREVIEW_ROW_LIMIT],
                'sheets': sheets,
                'warnings': extra_warnings,
                'content_type': 'table',
                'extraction_status': 'structured_extracted',
                'source_candidates': [
                    SourceCandidate(**candidate)
                    for candidate in build_source_candidates(
                        tables=[
                            {
                                'name': sheet.name,
                                'columns': sheet.columns,
                                'rows': sheet.rows,
                            }
                            for sheet in sheets
                        ]
                        or [{'name': original_name, 'columns': columns, 'rows': rows[:PREVIEW_ROW_LIMIT]}]
                    )
                ],
            }
        )
    elif ext in {'pdf', 'docx', 'txt'} | IMAGE_FILE_TYPES:
        parsed_kwargs.update(_parse_document(file_path))
    else:
        raise ParseError(f'Unsupported file type: {ext}')

    columns = list(parsed_kwargs['columns'])
    rows = list(parsed_kwargs['rows'])
    warnings = list(parsed_kwargs['warnings'])
    if not columns and not rows and not warnings:
        warnings.append('No columns detected in the file.')
    parsed_kwargs['warnings'] = warnings

    logger.info(
        'parse_file finished: name=%s ext=%s columns=%d preview_rows=%d warnings=%d content_type=%s extraction_status=%s',
        original_name,
        ext,
        len(columns),
        min(len(rows), PREVIEW_ROW_LIMIT),
        len(warnings),
        parsed_kwargs['content_type'],
        parsed_kwargs['extraction_status'],
    )

    return ParsedFile(
        file_name=original_name,
        file_type=ext,
        columns=columns,
        rows=rows[:PREVIEW_ROW_LIMIT],
        content_type=str(parsed_kwargs['content_type']),
        document_mode=str(parsed_kwargs['document_mode'] or 'data_table_mode'),
        extraction_status=str(parsed_kwargs['extraction_status']),
        raw_text=str(parsed_kwargs['raw_text']),
        text_blocks=list(parsed_kwargs['text_blocks']),
        sections=list(parsed_kwargs['sections']),
        kv_pairs=list(parsed_kwargs['kv_pairs']),
        source_candidates=list(parsed_kwargs['source_candidates']),
        sheets=list(parsed_kwargs['sheets']),
        form_model=parsed_kwargs.get('form_model'),
        pdf_zone_summary=dict(parsed_kwargs.get('pdf_zone_summary') or {}),
        warnings=warnings,
    )


def coerce_parsed_file(value: Any) -> ParsedFile:
    if isinstance(value, ParsedFile):
        return value
    if not isinstance(value, dict):
        raise ParseError('Parsed file payload must be an object.')

    text_blocks = [
        ParsedTextBlock(
            id=str(block.get('id') or f'block-{index}'),
            kind='line' if str(block.get('kind') or 'paragraph') == 'line' else 'paragraph',
            text=str(block.get('text') or ''),
            label=str(block.get('label')) if block.get('label') is not None else None,
        )
        for index, block in enumerate(value.get('text_blocks', []), start=1)
        if isinstance(block, dict)
    ]
    sections = [
        ParsedSection(title=str(section.get('title') or 'Section'), text=str(section.get('text') or ''))
        for section in value.get('sections', [])
        if isinstance(section, dict)
    ]
    kv_pairs = [
        ParsedKvPair(
            label=str(pair.get('label') or ''),
            value=str(pair.get('value') or ''),
            confidence=_safe_kv_confidence(pair.get('confidence')),
            source_text=str(pair.get('source_text')) if pair.get('source_text') is not None else None,
        )
        for pair in value.get('kv_pairs', [])
        if isinstance(pair, dict)
    ]
    source_candidates = [
        SourceCandidate(
            candidate_type=_safe_candidate_type(candidate.get('candidate_type')),
            label=str(candidate.get('label') or ''),
            value=candidate.get('value'),
            sample_values=list(candidate.get('sample_values') or []),
            source_text=str(candidate.get('source_text')) if candidate.get('source_text') is not None else None,
            section_title=str(candidate.get('section_title')) if candidate.get('section_title') is not None else None,
        )
        for candidate in value.get('source_candidates', [])
        if isinstance(candidate, dict)
    ]
    sheets = [
        ParsedSheet(
            name=str(sheet.get('name') or f'Sheet {index}'),
            columns=[str(column) for column in sheet.get('columns', [])],
            rows=[dict(row) for row in sheet.get('rows', []) if isinstance(row, dict)],
        )
        for index, sheet in enumerate(value.get('sheets', []), start=1)
        if isinstance(sheet, dict)
    ]
    form_model = _build_form_model(value.get('form_model'))
    return ParsedFile(
        file_name=str(value.get('file_name') or 'uploaded_file'),
        file_type=str(value.get('file_type') or 'unknown'),
        columns=[str(column) for column in value.get('columns', [])],
        rows=[dict(row) for row in value.get('rows', []) if isinstance(row, dict)],
        content_type=str(value.get('content_type') or 'unknown'),
        document_mode=str(value.get('document_mode') or 'data_table_mode'),
        extraction_status=str(value.get('extraction_status') or 'unknown'),
        raw_text=str(value.get('raw_text') or ''),
        text_blocks=text_blocks,
        sections=sections,
        kv_pairs=kv_pairs,
        source_candidates=source_candidates,
        sheets=sheets,
        form_model=form_model,
        pdf_zone_summary=dict(value.get('pdf_zone_summary') or {}),
        warnings=[str(warning) for warning in value.get('warnings', [])],
    )


def resolve_generation_source(
    parsed_file: ParsedFile,
    selected_sheet: str | None = None,
    target_fields: list[TargetField] | None = None,
) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    pdf_zone_routing = _get_pdf_zone_routing_signals(parsed_file)
    if parsed_file.form_model is not None and pdf_zone_routing:
        parsed_file.form_model.layout_meta['pdf_zone_routing'] = dict(pdf_zone_routing)

    if parsed_file.document_mode == 'form_layout_mode' and target_fields:
        if not pdf_zone_routing.get('prefer_table_source'):
            form_columns, form_rows, form_warnings = _resolve_form_layout_source(parsed_file, target_fields)
            if form_columns and form_rows:
                return form_columns, form_rows, form_warnings

    if parsed_file.document_mode == 'form_layout_mode' and not target_fields and _should_use_generic_form_source(parsed_file):
        form_columns, form_rows, form_warnings = _resolve_generic_form_layout_source(
            parsed_file,
            pdf_zone_routing=pdf_zone_routing,
        )
        if form_columns and form_rows:
            return form_columns, form_rows, form_warnings

    if not parsed_file.sheets:
        if parsed_file.columns or parsed_file.rows:
            warnings = []
            if pdf_zone_routing.get('prefer_table_source'):
                warnings.append('PDF zone routing preferred tabular extraction because table zones were stronger than form zones.')
            return parsed_file.columns, parsed_file.rows, warnings

        candidate_columns, candidate_rows = build_candidate_source(
            [_model_to_plain_dict(candidate) for candidate in parsed_file.source_candidates]
        )
        if candidate_columns and candidate_rows:
            return candidate_columns, candidate_rows, ['Generated mapping from extracted fields/text candidates.']

        if parsed_file.extraction_status in {'requires_ocr_or_manual_input', 'text_not_extracted', 'image_parse_not_supported_yet'}:
            raise ParseError(_extraction_status_message(parsed_file))

        return parsed_file.columns, parsed_file.rows, []

    if selected_sheet is None or selected_sheet.strip() == '':
        return parsed_file.columns, parsed_file.rows, []

    normalized_name = selected_sheet.strip()
    for sheet in parsed_file.sheets:
        if sheet.name == normalized_name:
            return (
                sheet.columns,
                sheet.rows,
                [f'Generated mapping from selected sheet: {sheet.name}'],
            )

    available_sheets = ', '.join(sheet.name for sheet in parsed_file.sheets)
    label = 'Worksheet' if parsed_file.file_type in {'xlsx', 'xls'} else 'Table'
    raise ParseError(f'{label} "{selected_sheet}" not found. Available: {available_sheets}')


def preview_business_form_resolutions(
    parsed_file: ParsedFile,
    *,
    target_fields: list[TargetField],
) -> list[FormFieldResolution]:
    if parsed_file.form_model is None:
        return []
    resolved_fields = resolve_business_form_fields(
        form_model=_model_to_plain_dict(parsed_file.form_model),
        target_fields=target_fields,
    )
    return [FormFieldResolution(**item) for item in resolved_fields]


def _resolve_form_layout_source(
    parsed_file: ParsedFile,
    target_fields: list[TargetField],
) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    if parsed_file.form_model is None:
        return [], [], []

    resolved_fields = resolve_business_form_fields(
        form_model=_model_to_plain_dict(parsed_file.form_model),
        target_fields=target_fields,
    )
    resolved_fields = _apply_form_pair_target_fallbacks(parsed_file, resolved_fields)
    parsed_file.form_model.resolved_fields = [FormFieldResolution(**item) for item in resolved_fields]
    parsed_file.form_model.layout_meta['requested_target_fields'] = [field.name for field in target_fields]
    pipeline_layers = dict(parsed_file.form_model.layout_meta.get('pipeline_layers') or {})
    pipeline_layers['business_mapping'] = {
        'status': 'completed',
        'requested_target_fields': [field.name for field in target_fields],
        'resolved_field_count': sum(
            1
            for item in resolved_fields
            if str(item.get('status') or '') in {'resolved', 'weak_match'} and item.get('value') is not None
        ),
    }
    parsed_file.form_model.layout_meta['pipeline_layers'] = pipeline_layers

    row: dict[str, Any] = {}
    columns: list[str] = []
    warnings: list[str] = ['Generated mapping from form-aware extraction.']
    critical_field_states: dict[str, str] = {}
    for item in resolved_fields:
        field_name = str(item.get('field') or '').strip()
        if not field_name:
            continue
        status = str(item.get('status') or 'not_found')
        normalized_field_name = field_name.casefold()
        if normalized_field_name in FORM_CRITICAL_TARGET_FIELDS:
            critical_field_states[field_name] = status
        value = item.get('value')
        if status in {'resolved', 'weak_match'} and value is not None:
            row[field_name] = value
            columns.append(field_name)
        elif status == 'ambiguous':
            warnings.append(
                f'Form field "{field_name}" is ambiguous and requires review.'
            )
        elif status == 'not_found':
            warnings.append(
                f'Form field "{field_name}" was not found in form-aware extraction.'
            )

    if columns:
        final_source_mode = 'repair_model' if any(
            field.resolved_by == 'repair_model' for field in parsed_file.form_model.resolved_fields
        ) else 'form_resolver'
        parsed_file.form_model.layout_meta['final_source_mode'] = final_source_mode
        parsed_file.form_model.layout_meta['resolved_columns'] = list(columns)
        parsed_file.form_model.layout_meta['critical_field_states'] = critical_field_states
        parsed_file.form_model.layout_meta['quality_summary'] = _assess_form_quality(
            parsed_file.form_model,
            target_fields=target_fields,
            resolved_fields=parsed_file.form_model.resolved_fields,
            resolved_columns=columns,
            final_source_mode=final_source_mode,
        )
        return columns, [row], _dedupe(warnings)

    unresolved_critical_fields = [
        field_name
        for field_name, status in critical_field_states.items()
        if status not in {'resolved', 'weak_match'}
    ]
    if unresolved_critical_fields:
        parsed_file.form_model.layout_meta['final_source_mode'] = 'fallback_blocked'
        parsed_file.form_model.layout_meta['critical_field_states'] = critical_field_states
        parsed_file.form_model.layout_meta['fallback_blocked_for_fields'] = list(unresolved_critical_fields)
        parsed_file.form_model.resolved_fields = [
            field if field.field.casefold() not in {name.casefold() for name in unresolved_critical_fields}
            else field.model_copy(update={'resolved_by': 'fallback_blocked'})
            if hasattr(field, 'model_copy')
            else FormFieldResolution(**{**_model_instance_to_dict(field), 'resolved_by': 'fallback_blocked'})
            for field in parsed_file.form_model.resolved_fields
        ]
        warnings.append(
            'Legacy fallback was blocked for critical form fields: '
            + ', '.join(unresolved_critical_fields)
            + '.'
        )
        parsed_file.form_model.layout_meta['quality_summary'] = _assess_form_quality(
            parsed_file.form_model,
            target_fields=target_fields,
            resolved_fields=parsed_file.form_model.resolved_fields,
            resolved_columns=[],
            final_source_mode='fallback_blocked',
        )
        return [], [], _dedupe(warnings)

    fallback_columns, fallback_rows = build_candidate_source(
        [_model_to_plain_dict(candidate) for candidate in parsed_file.source_candidates]
    )
    if fallback_columns and fallback_rows:
        parsed_file.form_model.layout_meta['final_source_mode'] = 'legacy_fallback'
        parsed_file.form_model.layout_meta['legacy_fallback_columns'] = list(fallback_columns)
        parsed_file.form_model.layout_meta['critical_field_states'] = critical_field_states
        parsed_file.form_model.layout_meta['quality_summary'] = _assess_form_quality(
            parsed_file.form_model,
            target_fields=target_fields,
            resolved_fields=parsed_file.form_model.resolved_fields,
            resolved_columns=fallback_columns,
            final_source_mode='legacy_fallback',
        )
        warnings.append('Form-aware extraction did not resolve target fields; fell back to extracted field/value candidates.')
        return fallback_columns, fallback_rows, _dedupe(warnings)

    parsed_file.form_model.layout_meta['final_source_mode'] = 'form_resolver'
    parsed_file.form_model.layout_meta['critical_field_states'] = critical_field_states
    parsed_file.form_model.layout_meta['quality_summary'] = _assess_form_quality(
        parsed_file.form_model,
        target_fields=target_fields,
        resolved_fields=parsed_file.form_model.resolved_fields,
        resolved_columns=[],
        final_source_mode='form_resolver',
    )
    return [], [], _dedupe(warnings)


def _resolve_generic_form_layout_source(
    parsed_file: ParsedFile,
    *,
    pdf_zone_routing: dict[str, Any] | None = None,
) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    if parsed_file.form_model is None:
        return [], [], []
    if pdf_zone_routing and pdf_zone_routing.get('prefer_table_source'):
        return [], [], ['PDF zone routing suppressed generic form source because table zones were stronger than form zones.']

    row: dict[str, Any] = {}
    columns: list[str] = []
    used_columns: set[str] = set()

    for scalar in parsed_file.form_model.scalars:
        label = str(scalar.label or '').strip()
        value = scalar.value
        if not _is_useful_form_label(label, value):
            continue
        unique_label = _make_unique_source_label(label, used_columns)
        row[unique_label] = value
        columns.append(unique_label)

    for group in parsed_file.form_model.groups:
        selected_labels = [
            str(option.label or '').strip()
            for option in group.options
            if option.selected and str(option.label or '').strip()
        ]
        if not selected_labels:
            continue

        label = _generic_form_group_label(group)
        if not label:
            continue

        unique_label = _make_unique_source_label(label, used_columns)
        row[unique_label] = selected_labels if group.group_type == 'multi_choice' or len(selected_labels) > 1 else selected_labels[0]
        columns.append(unique_label)

    if not columns:
        pair_columns, pair_row = _extract_form_pairs_from_rows(parsed_file)
        for label in pair_columns:
            unique_label = _make_unique_source_label(label, used_columns)
            row[unique_label] = pair_row.get(label)
            columns.append(unique_label)

    if columns:
        parsed_file.form_model.layout_meta['final_source_mode'] = 'generic_form_source'
        parsed_file.form_model.layout_meta['resolved_columns'] = list(columns)
        return columns, [row], ['Generated mapping from form-aware extracted fields.']

    fallback_columns, fallback_rows = build_candidate_source(
        [_model_to_plain_dict(candidate) for candidate in parsed_file.source_candidates]
    )
    if fallback_columns and fallback_rows:
        parsed_file.form_model.layout_meta['final_source_mode'] = 'legacy_fallback'
        parsed_file.form_model.layout_meta['legacy_fallback_columns'] = list(fallback_columns)
        return fallback_columns, fallback_rows, ['Generated mapping from extracted fields/text candidates.']

    return [], [], []


def _should_use_generic_form_source(parsed_file: ParsedFile) -> bool:
    if parsed_file.form_model is None:
        return False
    pdf_zone_routing = _get_pdf_zone_routing_signals(parsed_file)
    if pdf_zone_routing.get('prefer_table_source'):
        return False
    return any(any(option.selected for option in group.options) for group in parsed_file.form_model.groups)


def _generic_form_group_label(group: QuestionGroup) -> str:
    group_id = str(group.group_id or '').strip()
    if group_id and not _is_generic_form_group_id(group_id):
        return group_id
    return str(group.question or '').strip()


def _extract_form_pairs_from_rows(parsed_file: ParsedFile) -> tuple[list[str], dict[str, Any]]:
    if len(parsed_file.columns) < 2 or not parsed_file.rows:
        return [], {}

    label_column = parsed_file.columns[0]
    value_column = parsed_file.columns[1]
    row: dict[str, Any] = {}
    columns: list[str] = []
    consumed_group_rows, consumed_option_texts, consumed_question_texts = _collect_consumed_group_row_metadata(parsed_file)

    for row_index, source_row in enumerate(parsed_file.rows, start=1):
        label = str(source_row.get(label_column) or '').strip()
        value = source_row.get(value_column)
        if _is_consumed_group_pair_candidate(
            label=label,
            value=value,
            data_row_index=row_index,
            consumed_group_rows=consumed_group_rows,
            consumed_option_texts=consumed_option_texts,
            consumed_question_texts=consumed_question_texts,
        ):
            continue
        if not _is_useful_form_label(label, value):
            continue
        if label in row:
            continue
        row[label] = value
        columns.append(label)

    return columns, row


def _collect_consumed_group_row_metadata(
    parsed_file: ParsedFile,
) -> tuple[set[tuple[int, int]], set[str], set[str]]:
    if parsed_file.form_model is None:
        return set(), set(), set()

    consumed_group_rows: set[tuple[int, int]] = set()
    consumed_option_texts: set[str] = set()
    consumed_question_texts: set[str] = set()

    for group in parsed_file.form_model.groups:
        question_text = _normalize_form_text(group.question)
        if question_text:
            consumed_question_texts.add(question_text)

        source_ref = dict(group.source_ref or {})
        table_idx = source_ref.get('table_idx')
        row_idx = source_ref.get('row_idx')
        if isinstance(table_idx, int) and isinstance(row_idx, int):
            consumed_group_rows.add((table_idx, row_idx))

        for option in group.options:
            option_text = _normalize_form_text(option.label)
            if option_text:
                consumed_option_texts.add(option_text)

            option_source_ref = dict(option.source_ref or {})
            option_table_idx = option_source_ref.get('table_idx')
            option_row_idx = option_source_ref.get('row_idx')
            if isinstance(option_table_idx, int) and isinstance(option_row_idx, int):
                consumed_group_rows.add((option_table_idx, option_row_idx))

    return consumed_group_rows, consumed_option_texts, consumed_question_texts


def _is_consumed_group_pair_candidate(
    *,
    label: str,
    value: Any,
    data_row_index: int,
    consumed_group_rows: set[tuple[int, int]],
    consumed_option_texts: set[str],
    consumed_question_texts: set[str],
) -> bool:
    if (0, data_row_index) in consumed_group_rows:
        return True

    normalized_label = _normalize_form_text(label)
    normalized_value = _normalize_form_text(value)
    combined = ' '.join(part for part in [normalized_label, normalized_value] if part).strip()
    if not combined:
        return False

    if any(option_text and option_text in combined for option_text in consumed_option_texts):
        return True

    if _looks_like_option_marker_text(normalized_label) or _looks_like_option_marker_text(normalized_value):
        if any(question_text and question_text in combined for question_text in consumed_question_texts):
            return True

    return False


def _normalize_form_text(value: Any) -> str:
    return ' '.join(str(value or '').strip().casefold().split())


def _looks_like_option_marker_text(value: str) -> bool:
    normalized_value = str(value or '').strip().casefold()
    return normalized_value in {'x', '[x]', 'v', '[v]', '✓', '✔', '☒', '☑', 'да', 'нет', 'yes', 'no'}


def _is_useful_form_label(label: str, value: Any) -> bool:
    normalized_label = ' '.join(str(label or '').strip().split())
    if not normalized_label or value in (None, ''):
        return False
    if normalized_label.startswith(('•', '-', '*')):
        return False
    if _is_generic_form_group_id(normalized_label.casefold()):
        return False

    value_text = ' '.join(str(value).strip().split())
    if not value_text:
        return False

    label_letters = ''.join(ch for ch in normalized_label if ch.isalpha())
    value_letters = ''.join(ch for ch in value_text if ch.isalpha())
    if (
        len(label_letters) >= 16
        and len(value_letters) >= 8
        and label_letters.upper() == label_letters
        and value_letters.upper() == value_letters
    ):
        return False

    return True


def _is_generic_form_group_id(group_id: str) -> bool:
    normalized_group_id = str(group_id or '').strip().casefold()
    return not normalized_group_id or normalized_group_id in {'group', 'unknown'} or normalized_group_id.startswith('group_')


def _apply_form_pair_target_fallbacks(
    parsed_file: ParsedFile,
    resolved_fields: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not resolved_fields:
        return resolved_fields

    pair_columns, pair_row = _extract_form_pairs_from_rows(parsed_file)
    pair_labels = list(pair_columns)
    if not pair_row and parsed_file.form_model is None:
        return resolved_fields

    updated: list[dict[str, Any]] = []
    for item in resolved_fields:
        status = str(item.get('status') or 'not_found')
        if status in {'resolved', 'weak_match'} and item.get('value') is not None:
            updated.append(item)
            continue

        fallback = _resolve_target_field_from_form_pairs(
            field_name=str(item.get('field') or ''),
            pair_labels=pair_labels,
            pair_row=pair_row,
            parsed_file=parsed_file,
        )
        if fallback is None:
            updated.append(item)
            continue

        updated.append({**item, **fallback})

    return updated


def _resolve_target_field_from_form_pairs(
    *,
    field_name: str,
    pair_labels: list[str],
    pair_row: dict[str, Any],
    parsed_file: ParsedFile,
) -> dict[str, Any] | None:
    normalized_field_name = str(field_name or '').strip().casefold()
    if not normalized_field_name:
        return None

    if normalized_field_name == 'organizationname':
        match = _find_form_pair(pair_labels, pair_row, any_terms=('наименование организации', 'полное наименование организации', 'organization name', 'company name'))
        if match is not None:
            return _build_form_pair_resolution(match[0], match[1])
        return None

    if normalized_field_name in {'innorkio', 'innkio'}:
        match = _find_form_pair(pair_labels, pair_row, any_terms=('инн/кио', 'инн кио', 'инн', 'кио', 'inn', 'kio'))
        if match is not None:
            return _build_form_pair_resolution(match[0], match[1])
        return None

    if normalized_field_name in {'isresidentrf', 'istaxresidencyonlyrf', 'istaxresidentonlyinrussia'}:
        match = _find_form_pair(
            pair_labels,
            pair_row,
            all_terms=('выгодоприобрет', 'налогов', 'резидент'),
            exclude_terms=('контролир',),
        )
        if match is None:
            return None
        enum_value = _infer_tax_residency_enum(match[1])
        if enum_value is None:
            return None
        if normalized_field_name == 'isresidentrf':
            return _build_form_pair_resolution(match[0], enum_value, confidence=0.72)
        if normalized_field_name == 'istaxresidencyonlyrf':
            return _build_form_pair_resolution(match[0], enum_value == 'YES', confidence=0.72)
        return _build_form_pair_resolution(match[0], enum_value == 'YES', confidence=0.72)

    if normalized_field_name == 'arecontrollerstaxresidentsonlyinrussia':
        match = _find_form_pair(
            pair_labels,
            pair_row,
            all_terms=('контролир', 'налогов', 'резидент'),
        )
        if match is None:
            return None
        bool_value = _infer_yes_no_boolean(match[1])
        if bool_value is None:
            return None
        return _build_form_pair_resolution(match[0], bool_value, confidence=0.68)

    if normalized_field_name == 'anystatementtrue':
        bool_value = _infer_any_statement_true(parsed_file, pair_labels, pair_row)
        if bool_value is None:
            return None
        return _build_form_pair_resolution('fatca_statement_group', bool_value, confidence=0.7)

    if normalized_field_name == 'disregardedentityownertype':
        match = _find_form_pair(
            pair_labels,
            pair_row,
            any_terms=('disregarded entity является', 'disregarded entity owner', 'собственник owner disregarded entity является'),
        )
        if match is not None:
            return _build_form_pair_resolution(match[0], match[1], confidence=0.7)

        if parsed_file.form_model is not None:
            for group in parsed_file.form_model.groups:
                selected_labels = [
                    str(option.label or '').strip()
                    for option in group.options
                    if option.selected and str(option.label or '').strip()
                ]
                question = str(group.question or '').strip().casefold()
                if selected_labels and 'disregarded' in question:
                    return _build_form_pair_resolution(group.question, selected_labels[0], confidence=0.62)
        return None

    return None


def _find_form_pair(
    pair_labels: list[str],
    pair_row: dict[str, Any],
    *,
    any_terms: tuple[str, ...] = (),
    all_terms: tuple[str, ...] = (),
    exclude_terms: tuple[str, ...] = (),
) -> tuple[str, Any] | None:
    for label in pair_labels:
        normalized_label = _normalize_form_lookup(label)
        if any_terms and not any(_normalize_form_lookup(term) in normalized_label for term in any_terms):
            continue
        if all_terms and not all(_normalize_form_lookup(term) in normalized_label for term in all_terms):
            continue
        if exclude_terms and any(_normalize_form_lookup(term) in normalized_label for term in exclude_terms):
            continue
        return label, pair_row.get(label)
    return None


def _build_form_pair_resolution(label: str, value: Any, *, confidence: float = 0.74) -> dict[str, Any]:
    return {
        'status': 'weak_match',
        'resolved_by': 'legacy_fallback',
        'value': value,
        'candidates': [],
        'source_ref': {'source_type': 'row_pair', 'label': label},
        'confidence': confidence,
    }


def _infer_tax_residency_enum(value: Any) -> str | None:
    normalized_value = _normalize_form_lookup(value)
    if not normalized_value:
        return None
    if ' x не являюсь налоговым резидентом ни в одном государстве' in normalized_value:
        return 'NOWHERE'
    if normalized_value.startswith('да ') or normalized_value.startswith('да,') or ' x да ' in normalized_value:
        return 'YES'
    if normalized_value.startswith('нет ') or normalized_value.startswith('нет,') or ' x нет ' in normalized_value:
        return 'NO'
    if 'не являюсь налоговым резидентом ни в одном государстве' in normalized_value:
        return 'NOWHERE'
    return None


def _infer_yes_no_boolean(value: Any) -> bool | None:
    normalized_value = _normalize_form_lookup(value)
    if not normalized_value:
        return None
    if normalized_value.startswith('да ') or normalized_value.startswith('да,') or ' x да ' in normalized_value:
        return True
    if normalized_value.startswith('нет ') or normalized_value.startswith('нет,') or ' x нет ' in normalized_value:
        return False
    return None


def _infer_any_statement_true(
    parsed_file: ParsedFile,
    pair_labels: list[str],
    pair_row: dict[str, Any],
) -> bool | None:
    if parsed_file.form_model is not None:
        for group in parsed_file.form_model.groups:
            selected_labels = [
                str(option.label or '').strip()
                for option in group.options
                if option.selected and str(option.label or '').strip()
            ]
            if selected_labels:
                normalized_selected = ' '.join(_normalize_form_lookup(label) for label in selected_labels)
                if 'не применим' in normalized_selected:
                    return False
                return True

    match = _find_form_pair(pair_labels, pair_row, any_terms=('хотя бы одно из следующих утверждений', 'утверждений для выгодоприобретателя'))
    if match is None:
        return None

    normalized_value = _normalize_form_lookup(match[1])
    if not normalized_value:
        return None
    if 'не применим' in normalized_value and ' x ' not in normalized_value:
        return False
    return True


def _normalize_form_lookup(value: Any) -> str:
    text = str(value or '').casefold()
    collapsed = ' '.join(''.join(ch if ch.isalnum() else ' ' for ch in text).split())
    return collapsed


def _make_unique_source_label(label: str, used_columns: set[str]) -> str:
    if label not in used_columns:
        used_columns.add(label)
        return label

    suffix = 2
    while f'{label} {suffix}' in used_columns:
        suffix += 1
    unique_label = f'{label} {suffix}'
    used_columns.add(unique_label)
    return unique_label


def _model_instance_to_dict(value: Any) -> dict[str, Any]:
    if hasattr(value, 'model_dump'):
        return dict(value.model_dump())
    if hasattr(value, 'dict'):
        return dict(value.dict())
    if hasattr(value, '__dict__'):
        return dict(vars(value))
    return {}


def _parse_csv(file_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1251', 'latin-1']
    last_error: Exception | None = None

    for encoding in encodings_to_try:
        try:
            with file_path.open('r', encoding=encoding, newline='') as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                rows = [dict(row) for row in reader]
                columns = reader.fieldnames or []
                return [str(c) for c in columns], rows
        except Exception as exc:  # noqa: BLE001
            last_error = exc

    raise ParseError(f'Failed to parse CSV: {last_error}')



def _parse_excel(file_path: Path) -> tuple[list[str], list[dict[str, Any]], list[str], list[ParsedSheet]]:
    warnings: list[str] = []
    suffix = file_path.suffix.lower()

    try:
        if suffix == '.xlsx':
            return _parse_xlsx(file_path, warnings)
        if suffix == '.xls':
            return _parse_xls(file_path, warnings)
    except Exception as exc:  # noqa: BLE001
        raise ParseError(f'Failed to parse Excel: {exc}') from exc

    raise ParseError(f'Unsupported Excel file type: {suffix}')


def _parse_xlsx(file_path: Path, warnings: list[str]) -> tuple[list[str], list[dict[str, Any]], list[str], list[ParsedSheet]]:
    if load_workbook is None:
        raise ParseError('XLSX parsing is disabled in this build. Rebuild backend with openpyxl support.')

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        return _collect_excel_sheets(
            sheet_names=list(workbook.sheetnames),
            iter_sheet_rows=lambda sheet_name: workbook[sheet_name].iter_rows(values_only=True),
            warnings=warnings,
        )
    finally:
        workbook.close()


def _parse_xls(file_path: Path, warnings: list[str]) -> tuple[list[str], list[dict[str, Any]], list[str], list[ParsedSheet]]:
    if xlrd is None:
        raise ParseError('XLS parsing is disabled in this build. Rebuild backend with xlrd support or convert the file to XLSX.')

    workbook = xlrd.open_workbook(file_path)
    return _collect_excel_sheets(
        sheet_names=list(workbook.sheet_names()),
        iter_sheet_rows=lambda sheet_name: _iter_xls_sheet_rows(workbook, sheet_name),
        warnings=warnings,
    )


def _collect_excel_sheets(
    *,
    sheet_names: list[str],
    iter_sheet_rows,
    warnings: list[str],
) -> tuple[list[str], list[dict[str, Any]], list[str], list[ParsedSheet]]:
    combined_columns: list[str] = []
    combined_rows: list[dict[str, Any]] = []
    non_empty_sheets: list[str] = []
    sheets: list[ParsedSheet] = []

    for sheet_name in sheet_names:
        rows_iter = list(iter_sheet_rows(sheet_name))
        columns, rows, sheet_warnings = _sheet_rows_to_records(sheet_name, rows_iter)
        warnings.extend(sheet_warnings)

        if not columns and not rows:
            continue

        non_empty_sheets.append(sheet_name)
        sheets.append(ParsedSheet(name=sheet_name, columns=columns, rows=rows[:PREVIEW_ROW_LIMIT]))
        for column in columns:
            if column not in combined_columns:
                combined_columns.append(column)
        combined_rows.extend(rows)

    if len(non_empty_sheets) > 1:
        warnings.append(f'Merged {len(non_empty_sheets)} sheets: {", ".join(non_empty_sheets)}')
    elif len(sheet_names) > 1 and len(non_empty_sheets) == 1:
        warnings.append(f'Workbook has multiple sheets. Used the only non-empty sheet: {non_empty_sheets[0]}')

    return combined_columns, combined_rows[:PREVIEW_ROW_LIMIT], warnings, sheets


def _sheet_rows_to_records(sheet_name: str, rows_iter: list[tuple[Any, ...] | list[Any]]) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not rows_iter:
        return [], [], warnings

    raw_headers = list(rows_iter[0])
    if not raw_headers:
        return [], [], warnings

    if any(
        column is None
        or not isinstance(column, str)
        or (isinstance(column, str) and (column.strip() == '' or column.startswith('Unnamed:')))
        for column in raw_headers
    ):
        warnings.append(
            f'Sheet "{sheet_name}": Excel first row is treated as column headers. Some headers are empty or non-text, so the first row may contain data instead of column names.'
        )

    columns = [_stringify_excel_header(column, index) for index, column in enumerate(raw_headers, start=1)]
    rows: list[dict[str, Any]] = []

    for raw_row in rows_iter[1:]:
        values = list(raw_row)
        if not any(_has_visible_value(value) for value in values):
            continue
        record: dict[str, Any] = {}
        for index, column in enumerate(columns):
            value = values[index] if index < len(values) else None
            record[column] = _coerce_excel_value(value)
        rows.append(record)

    return columns, rows, warnings


def _iter_xls_sheet_rows(workbook: Any, sheet_name: str):
    sheet = workbook.sheet_by_name(sheet_name)
    for row_index in range(sheet.nrows):
        yield tuple(_coerce_xls_cell(workbook, sheet.cell(row_index, col_index)) for col_index in range(sheet.ncols))


def _coerce_xls_cell(workbook: Any, cell: Any) -> Any:
    if xlrd is None:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode).isoformat()
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        if float(cell.value).is_integer():
            return int(cell.value)
        return float(cell.value)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    return cell.value


def _stringify_excel_header(value: Any, index: int) -> str:
    if value is None:
        return f'Column {index}'
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or f'Column {index}'
    return str(value)


def _coerce_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _has_visible_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == '':
        return False
    return True


def _parse_document(file_path: Path) -> dict[str, Any]:
    try:
        parsed = parse_document(file_path)
        content_type = str(parsed.get('content_type', 'unknown'))
        extraction_status = str(parsed.get('extraction_status', 'unknown'))
        columns = [str(column) for column in parsed.get('columns', [])]
        raw_rows = parsed.get('rows', [])
        rows = [dict(row) for row in raw_rows if isinstance(row, dict)]
        warnings = [str(warning) for warning in parsed.get('warnings', [])]
        sheets: list[ParsedSheet] = []
        raw_tables = parsed.get('tables', [])
        if isinstance(raw_tables, list):
            for index, table in enumerate(raw_tables, start=1):
                if not isinstance(table, dict):
                    continue
                table_columns = [str(column) for column in table.get('columns', [])]
                table_rows = [dict(row) for row in table.get('rows', []) if isinstance(row, dict)]
                if not table_columns and not table_rows:
                    continue
                sheets.append(
                    ParsedSheet(
                        name=str(table.get('name') or f'Table {index}'),
                        columns=table_columns,
                        rows=table_rows[:PREVIEW_ROW_LIMIT],
                    )
                )
        original_row_count = len(rows)

        rows = _filter_sparse_rows(columns, rows)
        filtered_out = original_row_count - len(rows)
        if columns and not rows:
            warnings.append('Extracted table is too sparse to use as a structured source.')
            columns = []

        text_blocks = [
            ParsedTextBlock(
                id=str(block.get('id') or f'block-{index}'),
                kind='line' if str(block.get('kind') or 'paragraph') == 'line' else 'paragraph',
                text=str(block.get('text') or ''),
                label=str(block.get('label')) if block.get('label') is not None else None,
            )
            for index, block in enumerate(parsed.get('text_blocks', []), start=1)
            if isinstance(block, dict) and str(block.get('text') or '').strip()
        ]
        sections = [
            ParsedSection(title=str(section.get('title') or 'Section'), text=str(section.get('text') or ''))
            for section in parsed.get('sections', [])
            if isinstance(section, dict) and str(section.get('text') or '').strip()
        ]
        kv_pairs = [
            ParsedKvPair(
                label=str(pair.get('label') or ''),
                value=str(pair.get('value') or ''),
                confidence=_safe_kv_confidence(pair.get('confidence')),
                source_text=str(pair.get('source_text')) if pair.get('source_text') is not None else None,
            )
            for pair in parsed.get('kv_pairs', [])
            if isinstance(pair, dict) and str(pair.get('label') or '').strip() and str(pair.get('value') or '').strip()
        ]
        source_candidates = [
            SourceCandidate(
                candidate_type=_safe_candidate_type(candidate.get('candidate_type')),
                label=str(candidate.get('label') or ''),
                value=candidate.get('value'),
                sample_values=list(candidate.get('sample_values') or []),
                source_text=str(candidate.get('source_text')) if candidate.get('source_text') is not None else None,
                section_title=str(candidate.get('section_title')) if candidate.get('section_title') is not None else None,
            )
            for candidate in parsed.get('source_candidates', [])
            if isinstance(candidate, dict) and str(candidate.get('label') or '').strip()
        ]
        form_model_payload = parsed.get('form_model')
        form_model = _build_form_model(form_model_payload)

        if content_type not in {'table', 'mixed'} and not rows and not kv_pairs:
            warnings.append('Document uploaded without a tabular preview. Generation will use extracted fields when available.')

        logger.info(
            'document parser result: file=%s content_type=%s extraction_status=%s columns=%d rows=%d filtered_sparse_rows=%d kv_pairs=%d candidates=%d',
            file_path.name,
            content_type,
            extraction_status,
            len(columns),
            len(rows),
            max(filtered_out, 0),
            len(kv_pairs),
            len(source_candidates),
        )

        if content_type != 'table':
            logger.info(
                'document parser fallback to non-tabular mode: file=%s content_type=%s warnings=%s',
                file_path.name,
                content_type,
                warnings,
            )
        elif filtered_out > 0:
            logger.info(
                'document parser dropped sparse rows: file=%s dropped=%d remaining=%d',
                file_path.name,
                filtered_out,
                len(rows),
            )

        if sheets:
            filtered_sheets: list[ParsedSheet] = []
            for sheet in sheets:
                filtered_rows = _filter_sparse_rows(sheet.columns, [dict(row) for row in sheet.rows])
                if not sheet.columns and not filtered_rows:
                    continue
                filtered_sheets.append(
                    ParsedSheet(
                        name=sheet.name,
                        columns=sheet.columns,
                        rows=filtered_rows[:PREVIEW_ROW_LIMIT],
                    )
                )
            sheets = filtered_sheets

        return {
            'columns': columns,
            'rows': rows[:PREVIEW_ROW_LIMIT],
            'warnings': warnings,
            'sheets': sheets,
            'content_type': content_type,
            'document_mode': str(parsed.get('document_mode') or 'data_table_mode'),
            'extraction_status': extraction_status,
            'raw_text': str(parsed.get('text') or ''),
            'text_blocks': text_blocks,
            'sections': sections,
            'kv_pairs': kv_pairs,
            'source_candidates': source_candidates,
            'form_model': form_model,
            'pdf_zone_summary': dict(parsed.get('pdf_zone_summary') or {}),
        }
    except Exception as exc:  # noqa: BLE001
        logger.exception('document parser failed: file=%s error=%s', file_path.name, exc)
        raise ParseError(f'Failed to parse document: {exc}') from exc


def _filter_sparse_rows(columns: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not columns or not rows:
        return rows

    filtered_rows: list[dict[str, Any]] = []
    for row in rows:
        filled_cells = 0
        for column in columns:
            value = row.get(column)
            if value is None:
                continue
            if isinstance(value, str) and value.strip() == '':
                continue
            filled_cells += 1

        fill_ratio = filled_cells / max(len(columns), 1)
        if fill_ratio >= SPARSE_ROW_RATIO_THRESHOLD:
            filtered_rows.append(row)

    return filtered_rows


def _model_to_plain_dict(value: Any) -> dict[str, Any]:
    if hasattr(value, 'model_dump'):
        return _to_plain_jsonish(value.model_dump())
    if hasattr(value, 'dict'):
        return _to_plain_jsonish(value.dict())
    if isinstance(value, dict):
        return _to_plain_jsonish(value)
    raise TypeError(f'Unsupported parsed model type: {type(value)!r}')


def _to_plain_jsonish(value: Any) -> Any:
    if hasattr(value, 'model_dump'):
        return _to_plain_jsonish(value.model_dump())
    if hasattr(value, 'dict'):
        return _to_plain_jsonish(value.dict())
    if isinstance(value, dict):
        return {str(key): _to_plain_jsonish(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_plain_jsonish(item) for item in value]
    return value


def _build_form_model(value: Any) -> FormDocumentModel | None:
    if hasattr(value, 'model_dump'):
        value = value.model_dump()
    elif hasattr(value, 'dict'):
        value = value.dict()
    if not isinstance(value, dict):
        return None

    layout_lines = []
    for line in value.get('layout_lines', []):
        if hasattr(line, 'model_dump'):
            line = line.model_dump()
        elif hasattr(line, 'dict'):
            line = line.dict()
        if not isinstance(line, dict):
            continue
        tokens = []
        for token in line.get('tokens', []):
            if hasattr(token, 'model_dump'):
                token = token.model_dump()
            elif hasattr(token, 'dict'):
                token = token.dict()
            if isinstance(token, dict):
                tokens.append(LayoutToken(**token))
        layout_lines.append(LayoutLine(**{**line, 'tokens': tokens}))

    scalars = []
    for item in value.get('scalars', []):
        if hasattr(item, 'model_dump'):
            item = item.model_dump()
        elif hasattr(item, 'dict'):
            item = item.dict()
        if isinstance(item, dict):
            scalars.append(ScalarFieldCandidate(**item))
    groups = []
    for item in value.get('groups', []):
        if hasattr(item, 'model_dump'):
            item = item.model_dump()
        elif hasattr(item, 'dict'):
            item = item.dict()
        if not isinstance(item, dict):
            continue
        options = []
        for option in item.get('options', []):
            if hasattr(option, 'model_dump'):
                option = option.model_dump()
            elif hasattr(option, 'dict'):
                option = option.dict()
            if isinstance(option, dict):
                options.append(OptionItem(**option))
        groups.append(QuestionGroup(**{**item, 'options': options}))
    resolved_fields = []
    for item in value.get('resolved_fields', []):
        if hasattr(item, 'model_dump'):
            item = item.model_dump()
        elif hasattr(item, 'dict'):
            item = item.dict()
        if isinstance(item, dict):
            resolved_fields.append(FormFieldResolution(**item))

    return FormDocumentModel(
        scalars=scalars,
        groups=groups,
        structure=dict(value.get('structure') or {}),
        section_hierarchy=[dict(item) for item in value.get('section_hierarchy', []) if isinstance(item, dict)],
        layout_lines=layout_lines,
        layout_meta=dict(value.get('layout_meta') or {}),
        resolved_fields=resolved_fields,
    )


def _assess_form_quality(
    form_model: FormDocumentModel,
    *,
    target_fields: list[TargetField],
    resolved_fields: list[FormFieldResolution],
    resolved_columns: list[str],
    final_source_mode: str,
) -> dict[str, Any]:
    critical_field_states = dict(form_model.layout_meta.get('critical_field_states') or {})
    pdf_zone_routing = dict(form_model.layout_meta.get('pdf_zone_routing') or {})
    ambiguous_fields = [field.field for field in resolved_fields if field.status == 'ambiguous']
    unresolved_fields = [field.field for field in resolved_fields if field.status == 'not_found']
    repair_fields = [field.field for field in resolved_fields if field.resolved_by == 'repair_model']
    blocked_fields = [field.field for field in resolved_fields if field.resolved_by == 'fallback_blocked']
    unresolved_critical_fields = [
        field_name
        for field_name, status in critical_field_states.items()
        if status not in {'resolved', 'weak_match'}
    ]
    multiple_selected_single_choice_groups = [
        group.group_id
        for group in form_model.groups
        if group.group_type == 'single_choice' and sum(1 for option in group.options if option.selected) > 1
    ]

    red_flags: list[dict[str, Any]] = []
    if unresolved_critical_fields:
        red_flags.append(
            {
                'code': 'critical_unresolved',
                'message': 'Critical form fields were not reliably extracted.',
                'fields': unresolved_critical_fields,
            }
        )
    if ambiguous_fields:
        red_flags.append(
            {
                'code': 'ambiguous_fields',
                'message': 'Some form fields remain ambiguous.',
                'fields': ambiguous_fields,
            }
        )
    if multiple_selected_single_choice_groups:
        red_flags.append(
            {
                'code': 'single_choice_multi_select',
                'message': 'Single-choice groups contain multiple selected options.',
                'groups': multiple_selected_single_choice_groups,
            }
        )
    if final_source_mode == 'legacy_fallback':
        red_flags.append(
            {
                'code': 'legacy_fallback_used',
                'message': 'Legacy extracted candidates were used because form-aware extraction was incomplete.',
                'fields': list(resolved_columns),
            }
        )
    if not resolved_columns and target_fields:
        red_flags.append(
            {
                'code': 'empty_normalized_row',
                'message': 'Form-aware extraction produced an empty normalized row.',
            }
        )
    if target_fields and len(resolved_columns) <= max(len(target_fields) // 3, 1) and len(target_fields) >= 3:
        red_flags.append(
            {
                'code': 'low_field_coverage',
                'message': 'Only a small fraction of requested target fields was extracted.',
                'resolved_field_count': len(resolved_columns),
                'target_field_count': len(target_fields),
            }
        )
    if pdf_zone_routing.get('prefer_table_source'):
        red_flags.append(
            {
                'code': 'pdf_zone_prefers_table',
                'message': 'PDF table zones were stronger than form zones, so downstream should prefer tabular extraction.',
                'best_form_confidence': pdf_zone_routing.get('best_form_confidence'),
                'best_table_confidence': pdf_zone_routing.get('best_table_confidence'),
            }
        )
    elif pdf_zone_routing.get('has_form_zones') and pdf_zone_routing.get('low_confidence_form_zones'):
        red_flags.append(
            {
                'code': 'low_confidence_form_zones',
                'message': 'PDF form zones were detected with low confidence and should be reviewed before relying on them.',
                'best_form_confidence': pdf_zone_routing.get('best_form_confidence'),
            }
        )

    return {
        'needs_attention': bool(red_flags),
        'repair_recommended': bool(red_flags) and final_source_mode != 'repair_model',
        'resolved_field_count': len(resolved_columns),
        'target_field_count': len(target_fields),
        'scalar_candidate_count': len(form_model.scalars),
        'group_count': len(form_model.groups),
        'layout_line_count': len(form_model.layout_lines),
        'ambiguous_fields': ambiguous_fields,
        'unresolved_fields': unresolved_fields,
        'unresolved_critical_fields': unresolved_critical_fields,
        'repair_fields': repair_fields,
        'blocked_fields': blocked_fields,
        'multiple_selected_single_choice_groups': multiple_selected_single_choice_groups,
        'pdf_zone_routing': pdf_zone_routing,
        'red_flags': red_flags,
    }


def _get_pdf_zone_routing_signals(parsed_file: ParsedFile) -> dict[str, Any]:
    if str(parsed_file.file_type or '').casefold() != 'pdf':
        return {}

    zone_summary = dict(parsed_file.pdf_zone_summary or {})
    parser_outputs = dict(zone_summary.get('parser_outputs') or {})
    table_output = dict(parser_outputs.get('table') or {})
    form_output = dict(parser_outputs.get('form') or {})
    text_output = dict(parser_outputs.get('text') or {})
    noise_output = dict(parser_outputs.get('noise') or {})

    table_zone_confidences = _extract_zone_confidences(table_output.get('zones'))
    form_zone_confidences = _extract_zone_confidences(form_output.get('zones'))
    text_zone_confidences = _extract_zone_confidences(text_output.get('zones'))
    noise_zone_confidences = _extract_zone_confidences(noise_output.get('zones'))

    best_table_confidence = max(table_zone_confidences) if table_zone_confidences else None
    best_form_confidence = max(form_zone_confidences) if form_zone_confidences else None
    best_text_confidence = max(text_zone_confidences) if text_zone_confidences else None
    best_noise_confidence = max(noise_zone_confidences) if noise_zone_confidences else None

    has_confident_table_zone = best_table_confidence is not None and best_table_confidence >= PDF_TABLE_ZONE_MIN_CONFIDENCE
    has_confident_form_zone = best_form_confidence is not None and best_form_confidence >= PDF_FORM_ZONE_MIN_CONFIDENCE
    prefer_table_source = bool(
        has_confident_table_zone
        and parsed_file.columns
        and parsed_file.rows
        and (best_form_confidence is None or best_table_confidence > best_form_confidence + 0.05)
    )

    return {
        'available': bool(zone_summary),
        'dominant_zone': str(zone_summary.get('dominant_zone') or ''),
        'has_table_zones': bool(table_zone_confidences),
        'has_form_zones': bool(form_zone_confidences),
        'has_text_zones': bool(text_zone_confidences),
        'has_noise_zones': bool(noise_zone_confidences),
        'best_table_confidence': best_table_confidence,
        'best_form_confidence': best_form_confidence,
        'best_text_confidence': best_text_confidence,
        'best_noise_confidence': best_noise_confidence,
        'has_confident_table_zone': has_confident_table_zone,
        'has_confident_form_zone': has_confident_form_zone,
        'low_confidence_form_zones': bool(form_zone_confidences) and not has_confident_form_zone,
        'prefer_table_source': prefer_table_source,
    }


def _extract_zone_confidences(value: Any) -> list[float]:
    confidences: list[float] = []
    for item in list(value or []):
        if not isinstance(item, dict):
            continue
        zone_confidence = item.get('zone_confidence')
        try:
            if zone_confidence in (None, ''):
                continue
            confidences.append(float(zone_confidence))
        except (TypeError, ValueError):
            continue
    return confidences


def _safe_candidate_type(value: Any) -> str:
    candidate_type = str(value or 'text_section')
    if candidate_type in {'table_column', 'kv_pair', 'text_fact', 'text_section'}:
        return candidate_type
    return 'text_section'


def _safe_kv_confidence(value: Any) -> str:
    confidence = str(value or 'medium')
    if confidence in {'high', 'medium', 'low'}:
        return confidence
    return 'medium'


def _dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _extraction_status_message(parsed_file: ParsedFile) -> str:
    if parsed_file.extraction_status == 'requires_ocr_or_manual_input':
        return 'Document looks like a scan or image-based PDF. OCR or manual input is required.'
    if parsed_file.extraction_status == 'image_parse_not_supported_yet':
        return 'Image-like files are detected, but OCR-free extraction is not supported yet.'
    if parsed_file.extraction_status == 'text_not_extracted':
        return 'Text could not be extracted from the document.'
    return 'No usable source fields were extracted from the document.'



def infer_target_fields(target_json_raw: str) -> tuple[list[TargetField], dict[str, Any] | list[Any]]:
    target_fields, payload, _, _ = parse_target_schema(target_json_raw)
    return target_fields, payload


def parse_target_schema(
    target_json_raw: str,
) -> tuple[list[TargetField], dict[str, Any] | list[Any], dict[str, Any], dict[str, Any]]:
    payload, duplicate_keys = _load_target_json_with_duplicate_tracking(target_json_raw)

    if duplicate_keys:
        duplicates = ', '.join(sorted(set(duplicate_keys)))
        raise ParseError(f'Target JSON contains duplicate keys: {duplicates}')

    if not isinstance(payload, (dict, list)):
        raise ParseError('Target JSON must be an object or an array of objects.')

    target_schema = _build_target_schema(payload, path='$')
    root_object_schema = _resolve_root_object_schema(target_schema)
    target_fields = _extract_target_fields_from_schema(root_object_schema)
    schema_summary = {
        'root_type': target_schema['type'],
        'required_fields': list(root_object_schema.get('required_fields', [])),
        'field_count': len(target_fields),
        'root_is_array': target_schema['type'] == 'array',
    }
    return target_fields, payload, target_schema, schema_summary


def _load_target_json_with_duplicate_tracking(target_json_raw: str) -> tuple[Any, list[str]]:
    duplicate_keys: list[str] = []

    def _object_pairs_hook(pairs: list[tuple[Any, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            key_text = key if isinstance(key, str) else str(key)
            if key_text in result:
                duplicate_keys.append(key_text)
            result[key_text] = value
        return result

    try:
        payload = json.loads(target_json_raw, object_pairs_hook=_object_pairs_hook)
    except json.JSONDecodeError as exc:
        raise ParseError(f'Invalid target JSON: {exc}') from exc

    return payload, duplicate_keys


def _build_target_schema(value: Any, *, path: str) -> dict[str, Any]:
    if isinstance(value, dict):
        properties: dict[str, Any] = {}
        required_fields: list[str] = []
        for key, nested_value in value.items():
            if not isinstance(key, str):
                raise ParseError(f'Target JSON key at {path} must be a string.')
            normalized_key = key.strip()
            if not normalized_key:
                raise ParseError(f'Target JSON contains an empty key at {path}.')
            if normalized_key in properties:
                raise ParseError(f'Target JSON contains duplicate key "{normalized_key}" at {path}.')
            child_path = f'{path}.{normalized_key}'
            properties[normalized_key] = _build_target_schema(nested_value, path=child_path)
            required_fields.append(normalized_key)
        return {
            'type': 'object',
            'nullable': False,
            'properties': properties,
            'required_fields': required_fields,
        }

    if isinstance(value, list):
        if not value:
            return {
                'type': 'array',
                'nullable': False,
                'items': {'type': 'any', 'nullable': True},
            }

        item_schemas = [_build_target_schema(item, path=f'{path}[]') for item in value]
        merged_item_schema = _merge_array_item_schemas(item_schemas, path=f'{path}[]')
        return {
            'type': 'array',
            'nullable': False,
            'items': merged_item_schema,
        }

    inferred_type = _infer_type(value)
    return {
        'type': inferred_type,
        'nullable': value is None,
    }


def _merge_array_item_schemas(item_schemas: list[dict[str, Any]], *, path: str) -> dict[str, Any]:
    non_null_schemas = [schema for schema in item_schemas if schema.get('type') != 'null']
    nullable = len(non_null_schemas) != len(item_schemas)

    if not non_null_schemas:
        return {'type': 'null', 'nullable': True}

    merged = deepcopy(non_null_schemas[0])
    merged['nullable'] = bool(merged.get('nullable')) or nullable

    for schema in non_null_schemas[1:]:
        if merged.get('type') != schema.get('type'):
            raise ParseError(f'Target JSON contains conflicting array item types at {path}.')
        if merged.get('type') == 'object':
            merged = _merge_object_schemas(merged, schema, path=path)
            merged['nullable'] = bool(merged.get('nullable')) or nullable
            continue
        if merged.get('type') == 'array':
            merged['items'] = _merge_array_item_schemas(
                [merged.get('items', {'type': 'any'}), schema.get('items', {'type': 'any'})],
                path=f'{path}[]',
            )
            merged['nullable'] = bool(merged.get('nullable')) or nullable

    return merged


def _merge_object_schemas(left: dict[str, Any], right: dict[str, Any], *, path: str) -> dict[str, Any]:
    merged = deepcopy(left)
    left_properties = dict(merged.get('properties', {}))
    right_properties = dict(right.get('properties', {}))
    all_keys = sorted(set(left_properties) | set(right_properties))
    required_fields = set(merged.get('required_fields', [])) & set(right.get('required_fields', []))
    merged_properties: dict[str, Any] = {}

    for key in all_keys:
        child_path = f'{path}.{key}'
        left_schema = left_properties.get(key)
        right_schema = right_properties.get(key)
        if left_schema is None:
            merged_properties[key] = deepcopy(right_schema)
            continue
        if right_schema is None:
            merged_properties[key] = deepcopy(left_schema)
            continue
        if left_schema.get('type') != right_schema.get('type'):
            raise ParseError(f'Target JSON contains conflicting types for "{key}" at {child_path}.')
        if left_schema.get('type') == 'object':
            merged_properties[key] = _merge_object_schemas(left_schema, right_schema, path=child_path)
        elif left_schema.get('type') == 'array':
            merged_properties[key] = deepcopy(left_schema)
            merged_properties[key]['items'] = _merge_array_item_schemas(
                [left_schema.get('items', {'type': 'any'}), right_schema.get('items', {'type': 'any'})],
                path=f'{child_path}[]',
            )
        else:
            merged_properties[key] = deepcopy(left_schema)
            merged_properties[key]['nullable'] = bool(left_schema.get('nullable')) or bool(right_schema.get('nullable'))

    merged['properties'] = merged_properties
    merged['required_fields'] = [key for key in all_keys if key in required_fields]
    return merged


def _resolve_root_object_schema(target_schema: dict[str, Any]) -> dict[str, Any]:
    if target_schema.get('type') == 'object':
        return target_schema
    if target_schema.get('type') == 'array':
        item_schema = target_schema.get('items') or {}
        if item_schema.get('type') != 'object':
            raise ParseError('Target JSON array root must contain objects.')
        return item_schema
    raise ParseError('Target JSON must describe an object structure.')


def _extract_target_fields_from_schema(object_schema: dict[str, Any]) -> list[TargetField]:
    fields: list[TargetField] = []
    for key, field_schema in dict(object_schema.get('properties', {})).items():
        field_type = field_schema.get('type', 'any')
        if field_type not in {'string', 'number', 'boolean', 'object', 'array', 'null', 'any'}:
            field_type = 'any'
        fields.append(TargetField(name=key, type=field_type))
    return fields



def _infer_type(value: Any) -> str:
    if isinstance(value, bool):
        return 'boolean'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 'number'
    if isinstance(value, str):
        return 'string'
    if isinstance(value, list):
        return 'array'
    if isinstance(value, dict):
        return 'object'
    if value is None:
        return 'null'
    return 'any'
